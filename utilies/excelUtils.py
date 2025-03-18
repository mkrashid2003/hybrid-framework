
import openpyxl

def getRowCount(path,s1):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[s1]
    return sheet.max_row

def getColumnCount(path,s1):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[s1]
    return sheet.max_column

def readData(path,s1,r,c):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[s1]
    return sheet.cell(row=r, column=c).value


#def writeData(file,sheetName,rownum,columnno,data):
  #  workbook = openpyxl.load_workbook(file)
 #   sheet = workbook[sheetName]
 #   sheet.cell(row=rownum, column=columnno).value = data
 #   workbook.save(file)
